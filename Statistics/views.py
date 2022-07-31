from django.shortcuts import render
from openpyxl import load_workbook
from datetime import date as dt
from Friendship.settings import BASE_DIR


# Create your views here.

def response_stats(request):
    wb = load_workbook(str(BASE_DIR) + '/Экономика.xlsx', data_only=True)
    ws = wb.worksheets[2]
    config_ws = wb.worksheets[-1]

    names = [(ws.cell(row, 2).value, row) for row in range(3, 7)]
    money = [
        [
            [name[0][3:], ws.cell(name[1], i).value, ws.cell(name[1], i).column]
            for i in range(3, 16) if ws.cell(name[1], i).value is not None][-1]
        for name in names
    ]
    config_info = [
        [config_ws.cell(row, col).value for col in range(1, 4)]
        for row in range(1, 5)
    ]

    friendship = dict(
        (
            name[0][3:],
            [
                config_info[pos][1], money[pos][1], config_info[pos][-1]
            ]
        ) for pos, name in enumerate(names)
    )

    actual_date = ws.cell(2, money[0][-1]).value

    games = [
        [
            ws.cell(row, 1).value, ws.cell(row, 2).value,
            [
                (col - 2, ws.cell(row, col).value) for col in range(3, 7)
            ]
        ]
        for row in range(13, 30) if ws.cell(row, 1).value is not None and ws.cell(row, 3).value != '-'
    ]

    context = {
        "friendships": friendship,
        "games": games,
        "today": actual_date,
    }

    return render(request, 'statistics/Stats.html', context=context)
